# Chpater 13
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames) # The workbook's sheets' names.
sheet = wb['Sheet1'] # Get a sheet from the workbook.
print(sheet)
print(type(sheet))
print(sheet.title) # Get the sheet's title as a string.
anotherSheet = wb.active # Get the active sheet.
print(anotherSheet)

print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
# Get the row, column, and value from the cell.
print('Row %s, Colum %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print(sheet['C1'].value)
print(sheet.cell(row=1,column=2))
print(sheet.cell(row=1,column=2).value)
for i in range(1,8,2): # Go through every onter row:
    print(i, sheet.cell(row=i,column=2).value)
print(sheet.max_row) # Get the hightest row number.
print(sheet.max_column)
